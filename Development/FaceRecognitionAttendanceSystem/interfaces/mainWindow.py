from tkinter import BASELINE
from interfaces import global_initialize as global_ini

import os
if os.name=='nt':
    import winsound

import requests
import json
import sys
import cv2
import pandas as pd
import re
import time
from datetime import datetime
import qimage2ndarray
from threading import Thread
from src.commons import functions, distance as dst
from src.detectors import FaceDetector

from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.uic import loadUi

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
import numpy as np
import PIL.Image

text_color = (255,255,255) # white color

time_threshold = 5
frame_threshold = 10
pivot_img_size = 112 #face recognition result image
freeze = False
face_detected = False
face_included_frames = 0 #freeze screen if face detected sequantially 5 frames
freezed_frame = 0
tic = time.time()

screensaver = False

sound_counter = 0
verified = False

# ---------------------------------------Main Windows--------------------------------------

class mainWindow(QDialog):
    # Run One Time
    df = pd.read_pickle("trainer/face_embeddings.pkl")
    uploadTime = time.perf_counter()
    latestTime = time.perf_counter()
    namelist = [] 
    
    staticTime = time.perf_counter()

    ret, frame1 = global_ini.cap.read()
    ret, frame2 = global_ini.cap.read()

    def __init__(self):
        
        super(mainWindow, self).__init__()
        loadUi("ui/main.ui", self)
        self.setWindowIcon(QIcon('faceid.ico'))
        self.df = pd.read_pickle("trainer/face_embeddings.pkl")
        print("Embeddings loaded")
        timer = QTimer(self)
        timer.timeout.connect(self.displayFrame)
        timer.start(1)
     
#-------------------------- obtain face detections--------------------------        
    
    def displayFrame(self):
        #------------------------ 
        global freeze
        global face_detected
        global face_included_frames
        global freezed_frame
        global tic
        global freeze_img
        global time_threshold
        global frame_threshold
        global sound_counter
        global verified
        global screensaver
 
        #------------------------
        
        #================= Create a list to avoid taking duplicate attendance in shorter time ===============
        uploadBufferTime = (time.perf_counter()) - mainWindow.uploadTime    
        if uploadBufferTime >= 60 and len(mainWindow.namelist) > 0:
            mainWindow.uploadTime = time.perf_counter()
            mainWindow.namelist.pop(0)
            print("Upload Buffer Time: " + str(uploadBufferTime))
            print(mainWindow.namelist)
        elif uploadBufferTime >= 30 and len(mainWindow.namelist) == 0:
            mainWindow.uploadTime = time.perf_counter()
        #================= Create a list to avoid taking duplicate attendance in shorter time ===============

        #------------------------
    # while(True):
        # ret, img = global_ini.cap.read()
        # img = cv2.flip(img,1)

        # ret, frame1 = global_ini.cap.read()
        # ret, frame2 = global_ini.cap.read()

        img = cv2.flip(self.frame1,1)
        img2 = cv2.flip(self.frame2,1)

        diff = cv2.absdiff(img, img2) # calculate absolute different between current and previous pixel
        gray = cv2.cvtColor(diff, cv2.COLOR_BGR2GRAY) # convert to gray scale
        blur = cv2.GaussianBlur(gray, (5,5), 0) # 
        _, thresh = cv2.threshold(blur, 30, 255, cv2.THRESH_BINARY) #Adjust threshold here (eg. all pixel above 30 will convert to 255(white pixel))
        dilated = cv2.dilate(thresh, None, iterations=3)
        contours, _ = cv2.findContours(dilated, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

        for contour in contours:
            (x, y, w, h) = cv2.boundingRect(contour)

            if cv2.contourArea(contour) < 900: # control how big the movement you want to detect
                if (time.perf_counter() - self.staticTime) >= 10: #现在minus静止时间 more than 10sec -> ScreenSaver
                    screensaver = True

                # no movement
                continue
            # cv2.drawContours(frame1, contours, -1, (0, 255, 0), 2)
            # cv2.rectangle(img, (x, y), (x+w, y+h), (0, 255, 0), 2)
            # cv2.putText(img, "Status: {}".format('Movement'), (10, 20), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 3)
            self.staticTime = time.perf_counter()
            screensaver = False

        # display latest frame
        # img = img2

        if img is None:
            return

        #cv2.namedWindow('img', cv2.WINDOW_FREERATIO)
        #cv2.setWindowProperty('img', cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)

        raw_img = img.copy()
        resolution = img.shape; resolution_x = img.shape[1]; resolution_y = img.shape[0]

        if freeze == False:

            try:
                #faces store list of detected_face and region pair
                faces = FaceDetector.detect_faces(global_ini.face_detector, global_ini.detector_backend, img, align = True)
            except: #to avoid exception if no face detected
                faces = []

            if len(faces) == 0:
                face_included_frames = 0
        else:
            faces = []

        detected_faces = []
        face_index = 0
        for face, (x, y, w, h) in faces:
            if w > 130: #discard small detected faces

                face_detected = True
                if face_index == 0:
                    face_included_frames = face_included_frames + 1 #increase frame for a single face

                cv2.rectangle(img, (x,y), (x+w,y+h), (67,67,67), 2) #draw rectangle to main image

                cv2.putText(img, str(frame_threshold - face_included_frames), (int(x+w/2.5),int(y+h/1.5)), cv2.FONT_HERSHEY_SIMPLEX, 3, (255, 255, 255), 2)

                detected_face = img[int(y):int(y+h), int(x):int(x+w)] #crop detected face

                #-------------------------------------
                detected_faces.append((x,y,w,h))
                face_index = face_index + 1

                screensaver = False
                #-------------------------------------

        if face_detected == True and face_included_frames == frame_threshold and freeze == False:
            freeze = True
            #base_img = img.copy()
            base_img = raw_img.copy()
            detected_faces_final = detected_faces.copy()
            tic = time.time()

        if freeze == True:

            toc = time.time()
            if (toc - tic) < time_threshold:

                if freezed_frame == 0:
                    freeze_img = base_img.copy()
                    #freeze_img = np.zeros(resolution, np.uint8) #here, np.uint8 handles showing white area issue

                    for detected_face in detected_faces_final:
                        x = detected_face[0]; y = detected_face[1]
                        w = detected_face[2]; h = detected_face[3]

                        # freeze 着的时候画的image
                        cv2.rectangle(freeze_img, (x,y), (x+w,y+h), (50,205,50), 1) #draw rectangle to main image

                        #-------------------------------

                        #apply deep learning for custom_face

                        custom_face = base_img[y:y+h, x:x+w]

                        #-------------------------------
                        #face recognition

                        custom_face = functions.preprocess_face(img = custom_face, target_size = (global_ini.input_shape_y, global_ini.input_shape_x), enforce_detection = False, detector_backend = global_ini.detector_backend)

                        #check preprocess_face function handled
                        if custom_face.shape[1:3] == global_ini.input_shape:
                            if self.df.shape[0] > 0: #if there are images to verify, apply face recognition
                                img1_representation = global_ini.model.predict(custom_face)[0,:]

                                #print(freezed_frame," - ",img1_representation[0:5])

                                def findDistance(row):
                                    distance_metric = row['distance_metric']
                                    img2_representation = row['embedding']

                                    distance = 1000 #initialize very large value
                                    if distance_metric == 'cosine':
                                        distance = dst.findCosineDistance(img1_representation, img2_representation)
                                    elif distance_metric == 'euclidean':
                                        distance = dst.findEuclideanDistance(img1_representation, img2_representation)
                                    elif distance_metric == 'euclidean_l2':
                                        distance = dst.findEuclideanDistance(dst.l2_normalize(img1_representation), dst.l2_normalize(img2_representation))

                                    return distance
                                # ------------ Play Sound Effect ------------- 
                                # Only for Linux System (mpg321)
                                def playEffect(sound_counter, verified):
                                    if sound_counter == 1 and verified == True:
                                        if os.name=='posix':
                                            os.system("mpg321 sound/success.mp3")
                                        elif os.name=='nt':
                                            winsound.PlaySound("sound/success.wav", winsound.SND_FILENAME )
                                        sound_counter += 1
                                    elif sound_counter == 1 and verified == False:
                                        if os.name=='posix':
                                            os.system("mpg321 sound/fail.mp3")
                                        elif os.name=='nt':
                                            winsound.PlaySound("sound/fail.wav", winsound.SND_FILENAME )
                                        sound_counter += 1
                                        
                                self.df['distance'] = self.df.apply(findDistance, axis = 1)
                                self.df = self.df.sort_values(by = ["distance"])

                                candidate = self.df.iloc[0]
                                employee_name = candidate['employee']
                                best_distance = candidate['distance']

                                #print(candidate[['employee', 'distance']].values)

                                #if True:
                                print(global_ini.distance_metric,"Best Distance:",best_distance)
                                print("Predicted Name:",employee_name)
                                print("Threshold: ",global_ini.threshold)
                                print("* Confidence Score: {:.2%} \n-------------------".format(1-best_distance))

                                # ===============================     LOGGING      =================================
                                if global_ini.checkBox_logging:
                                    if os.path.exists("logging.xlsx"):
                                        wb = load_workbook(filename = 'logging.xlsx')
                                        ws = wb.worksheets[0]
                                        num = ws.max_row

                                        cv2.imwrite('logging' + '/realTime' + "-" + str(num) + ".png", base_img[y:y+h, x:x+w])
                                        img = Image('logging/realTime-{}.png'.format(num))

                                        img.height=112
                                        img.width=112
                                        ws.row_dimensions[num+1].height=88
                                        ws.add_image(img, anchor = 'C'+str(num+1))

                                        img2 = Image(employee_name)
                                        img2.height=112
                                        img2.width=112
                                        ws.add_image(img2, anchor='G'+str(num+1))

                                        excel_label = employee_name.split("/")[-1].replace(".jpg", "")
                                        excel_label = re.sub('-[0-9]', '', excel_label)
                                        excel_dateStr = QDateTime.currentDateTime().toString("yyyy-MM-dd hh:mm:ss")
                                        
                                        ws['A'+str(num+1)] = num
                                        ws['B'+str(num+1)] = excel_dateStr
                                        ws['D'+str(num+1)] = excel_label
                                        ws['E'+str(num+1)] = "{:.2%}".format(1-best_distance)
                                        ws['F'+str(num+1)] = global_ini.model_name+": "+str(global_ini.threshold)
                                        ws['H'+str(num+1)] = employee_name

                                        if best_distance <= float(global_ini.threshold):
                                            ws['I'+str(num+1)] = "Success"
                                        else:
                                            ws['I'+str(num+1)] = "Fail"

                                        wb.save('logging.xlsx')
                                    else:
                                        wb = Workbook()
                                        ws = wb.worksheets[0]

                                        ws['A1'] = "No."
                                        ws['B1'] = "DateTime"
                                        ws['C1'] = "Face Captured"
                                        ws['D1'] = "Predicted Name"
                                        ws['E1'] = "Confidence Score"
                                        ws['F1'] = "Model & Threshold"
                                        ws['G1'] = "Predicted Image"
                                        ws['H1'] = "Predicted Path"
                                        ws['I1'] = "Result"
                                        
                                        cv2.imwrite('logging' + '/realTime' + "-" + str(1) + ".png", base_img[y:y+h, x:x+w])
                                        img = Image('logging/realTime-1.png')

                                        img.height=112
                                        img.width=112
                                        ws.row_dimensions[2].height=88
                                        ws.column_dimensions['A'].width=16 #No
                                        ws.column_dimensions['B'].width=22 #date
                                        ws.column_dimensions['C'].width=16 #face
                                        ws.column_dimensions['D'].width=20 #pred name
                                        ws.column_dimensions['E'].width=16 #conf
                                        ws.column_dimensions['F'].width=20 #model
                                        ws.column_dimensions['G'].width=16 #pred img
                                        ws.column_dimensions['H'].width=30 #pred path
                                        ws.column_dimensions['I'].width=10 #result
                                        ws.add_image(img, anchor ='C2')

                                        img2 = Image(employee_name)
                                        img2.height=112
                                        img2.width=112
                                        ws.add_image(img2, anchor='G2')

                                        excel_label = employee_name.split("/")[-1].replace(".jpg", "")
                                        excel_label = re.sub('-[0-9]', '', excel_label)
                                        excel_dateStr = QDateTime.currentDateTime().toString("yyyy-MM-dd hh:mm:ss")
                                        
                                        ws['A2'] = 1
                                        ws['B2'] = excel_dateStr
                                        ws['D2'] = excel_label
                                        ws['E2'] = "{:.2%}".format(1-best_distance)
                                        ws['F2'] = global_ini.model_name+": "+str(global_ini.threshold)
                                        ws['H2'] = employee_name


                                        if best_distance <= float(global_ini.threshold):
                                            ws['I2'] = "Success"
                                        else:
                                            ws['I2'] = "Fail"

                                        wb.save('logging.xlsx')
                                # ==================================================================================

                                if best_distance <= float(global_ini.threshold):
                                    
                                    time_threshold = 4
                                    predicted_name = re.sub(r'-\d+\.jpg', '-1.jpg', employee_name) # Display the same comparable image | eg. display JH-1.jpg instead of JH-2.jpg & JH-3.jpg
                                    #print(employee_name)
                                    print(predicted_name)
                                    display_img = cv2.imread(predicted_name)
                                    
                                    #display_img 原本对比的图像
                                    detect_face_display_img = FaceDetector.detect_faces(global_ini.face_detector, global_ini.detector_backend, display_img, align = True)
                                    
                                    for face, (a, b, c, d) in detect_face_display_img:
                                        display_img = display_img[int(b):int(b+d), int(a):int(a+c)]
                                    
                                    try:
                                        display_img = cv2.resize(display_img, (pivot_img_size, pivot_img_size))
                                    except:
                                        print("Fail")
                                    label = employee_name.split("/")[-1].replace(".jpg", "")
                                    label = re.sub('-[0-9]', '', label)
                                    # original : GANJH-1.jpg / GANJH-2.jpy -> GANJH

                                    # =====================   Joo Han   ====================================
                                    # Query data from MySQL
                                    sql = "SELECT user_id FROM employee WHERE user_name = %s"
                                    val = (label, )

                                    try:
                                        global_ini.mycursor.execute(sql, val)
                                    except Exception as e:
                                        QMessageBox.critical(None, "Error", str(repr(e)))
                                        return
                                    
                                    output = global_ini.mycursor.fetchone()
                                    for n in output:
                                        userid = n

                                    #dateStr = ""
                                    dateNow = QDate.currentDate().toString(Qt.ISODate)
                                    timeNow = QTime.currentTime().toString() #Qt.DefaultLocaleLongDate
                                    #dateStr = dateNow + "  " + timeNow
                                    dateStr = QDateTime.currentDateTime().toString("yyyy-MM-dd hh:mm:ss")
                                    self.IDLabel.setText(str(userid))
                                    self.Namelabel.setText(str(label))
                                    self.Timelabel.setText(dateStr)
                                    
                                    # ===
                                    bufferTime = (time.perf_counter()) - mainWindow.latestTime
                                    if bufferTime >= 3:
                                    
                                        if label not in mainWindow.namelist:
                                            mainWindow.namelist.append(label)
                                            mainWindow.latestTime = time.perf_counter()
                                            print("Buffer Time: " + str(bufferTime))
                                            print("Attendance Taken: "+label+" Date: "+dateNow+" Time: "+timeNow)
                                            sql = "INSERT INTO attendance (user_id, user_name, sign_in, machine) VALUES (%s, %s, %s, %s)"
                                            val = (userid, label, dateStr, global_ini.machine_code)
                                            global_ini.mycursor.execute(sql, val)
                                            global_ini.mydb.commit()
                                            self.messagelabel.setText("Successfully verified. Attendance is taken ✓")
                                            sound_counter += 1
                                            verified = True
                                            thread = Thread(target = playEffect, args = (sound_counter,verified))
                                            thread.start()
                                            # continue

                                            # Webhook
                                            try:
                                                sql = "SELECT url, status, event FROM webhook_table"
                                                try:
                                                    global_ini.mycursor.execute(sql,)
                                                except Exception as e:
                                                    QMessageBox.critical(None, "Error", str(repr(e)))
                                                    return

                                                webhook_output = global_ini.mycursor.fetchall()

                                                for row in webhook_output:
                                                    if row[1]=="Enabled" and row[2]=="Attendance":

                                                        try:
                                                            data = {'Event':row[2], 'OID':userid, 'Name':label, "Attendance":dateStr, "Machine No.":global_ini.machine_code}
                                                            r = requests.post(row[0], data=json.dumps(data), headers={'Content-Type':'application/json'})
                                                        except:
                                                            pass

                                            except Exception as e:
                                                print(e)
                                                
                                        
                                        else:
                                            self.messagelabel.setText("Your attendance has been recorded. ")
                                            sound_counter += 1
                                            verified = True
                                            thread = Thread(target = playEffect, args = (sound_counter,verified))
                                            thread.start()                                            
                                            # continue 
                                    # ======================= Joo Han ==========================================
                                    try:
                                        x = 179; y=57; w=264; h=264
                                        freeze_img[y+h:y+h+pivot_img_size, x-pivot_img_size:x] = display_img

                                        overlay = freeze_img.copy(); opacity = 0.2

                                        # border
                                        cv2.rectangle(freeze_img, (x-pivot_img_size, y+h), (x,y+h+pivot_img_size), (46,200,255), 3)
                                        # Finds space required by the text so that we can put a background with that amount of width.
                                        picture_label = "Verified: "+label
                                        (z, p), baseline = cv2.getTextSize(picture_label, cv2.FONT_HERSHEY_DUPLEX, 0.6, 1)

                                        cv2.rectangle(freeze_img,(x+2, y+h+pivot_img_size-25),(x+z+5, y+h+pivot_img_size+2),(46,200,255),cv2.FILLED)
                                        cv2.addWeighted(overlay, opacity, freeze_img, 1 - opacity, 0, freeze_img)

                                        cv2.putText(freeze_img, picture_label, (x+3, y+h+pivot_img_size-5), cv2.FONT_HERSHEY_DUPLEX, 0.6, text_color, 1)
                                       
                                    except Exception as err:
                                        print(str(err))
                                
                                else:
                                    time_threshold = 3
                                    cv2.rectangle(freeze_img, (x,y), (x+w,y+h), (0,0,255), 2)
                                    self.IDLabel.setText("-")
                                    self.Namelabel.setText("unknown")
                                    self.Timelabel.setText("-")
                                    self.messagelabel.setText("Warning !!! Unknown Person...")
                                    
                                    sound_counter += 1
                                    verified = False
                                    thread = Thread(target = playEffect, args = (sound_counter,verified))
                                    thread.start()                                    

                        tic = time.time() #in this way, freezed image can show 5 seconds

              
                # ======== if freeze != 0 ==========

                time_left = int(time_threshold - (toc - tic) + 1)

                cv2.rectangle(freeze_img, (10, 10), (90, 50), (67,67,67), -10)
                cv2.putText(freeze_img, str(time_left), (40, 40), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 1)
                    
                # cv2.imshow('img', freeze_img)
                freeze_img2 = cv2.cvtColor(freeze_img, cv2.COLOR_BGR2RGB)
                freeze_img2 = qimage2ndarray.array2qimage(freeze_img2)
                self.videolabel.setPixmap(QPixmap.fromImage(freeze_img2))
                
                
                freezed_frame = freezed_frame + 1
                
                self.frame1 = self.frame2
                ret, self.frame2 = global_ini.cap.read()
                # ===================================

            else:
                face_detected = False
                face_included_frames = 0
                freeze = False
                freezed_frame = 0

                verified = False
                sound_counter = 0
                
                self.Namelabel.clear() 
                self.IDLabel.clear()
                self.Timelabel.clear()
                self.messagelabel.clear()

        else:

            if screensaver:
                self.label_clock.setVisible(True)
                self.label_clock_2.setVisible(True)
                img = cv2.imread('images/clock2.jpg')
                dt_string = datetime.now().strftime("%H:%M:%S")
                dt_string2 = datetime.now().strftime("%d %B, %Y")
                self.label_clock.setText(dt_string)
                self.label_clock_2.setText(dt_string2)
                img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                img = qimage2ndarray.array2qimage(img)
                self.videolabel.setPixmap(QPixmap.fromImage(img))
                #'%A %d %B %Y %I:%M:%S%p' (Friday 21 Octorber 2022 02:12:38PM)
            else:
                self.label_clock.setVisible(False)
                self.label_clock_2.setVisible(False)
                img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                img = qimage2ndarray.array2qimage(img)
                self.videolabel.setPixmap(QPixmap.fromImage(img))

            self.frame1 = self.frame2
            ret, self.frame2 = global_ini.cap.read()